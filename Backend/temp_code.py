import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import logging
import shutil
import os

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def process_workbook(filename, sheet_name=None, adjustment_factor=0.9, create_backup=True):
    """
    Processes an Excel workbook to adjust prices in a specific column
    and adds a bar chart based on the corrected prices.

    Args:
    - filename (str): The path to the Excel file.
    - sheet_name (str, optional): The name of the sheet to process (default is None, which will use the first sheet).
    - adjustment_factor (float, optional): The factor by which to adjust prices (default is 0.9).
    - create_backup (bool, optional): Whether to create a backup of the workbook before saving (default is True).
    """
    try:
        logging.info(f"Starting to process workbook: {filename}")

        wb = xl.load_workbook(filename)
        
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                logging.error(f"Sheet '{sheet_name}' not found in the workbook.")
                return
            sheet = wb[sheet_name]
        else:
            sheet = wb.active
            logging.info(f"No sheet name provided, using the first sheet: '{sheet.title}'")
        
        # Create backup if specified
        if create_backup:
            backup_filename = f"{filename}.backup"
            shutil.copyfile(filename, backup_filename)
            logging.info(f"Backup created at '{backup_filename}'")
        
        changes_made = False
        
        # Adjust prices and insert corrected prices in column 4
        for row in range(2, sheet.max_row + 1):
            price_cell = sheet.cell(row, 3)  # Column 3 holds original prices
            if isinstance(price_cell.value, (int, float)):
                corrected_price = price_cell.value * adjustment_factor
                corrected_price_cell = sheet.cell(row, 4)  # Column 4 holds corrected prices
                corrected_price_cell.value = corrected_price
                changes_made = True
            else:
                logging.warning(f"Non-numeric or empty value in row {row}, column 3: {price_cell.value}")

        # Add a bar chart for corrected prices if data exists
        if sheet.max_row > 1:
            values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
            chart = BarChart()
            chart.add_data(values, titles_from_data=False)
            chart.title = "Corrected Prices Chart"
            chart.x_axis.title = "Product"
            chart.y_axis.title = "Corrected Price"
            sheet.add_chart(chart, 'G2')  
            logging.info(f"Chart added to the sheet '{sheet.title}'.")
        else:
            logging.warning("No data to create a chart.")
        
        # Save only if there are changes
        if changes_made:
            wb.save(filename)
            logging.info(f"Workbook '{filename}' processed successfully and saved.")
        else:
            logging.info(f"No changes made to the workbook '{filename}', so it was not saved.")

    except FileNotFoundError:
        logging.error(f"Error: File '{filename}' not found.")
    except Exception as e:
        logging.error(f"An error occurred: {e}")

# Example usage
process_workbook('transactions.xlsx')